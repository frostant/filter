# contestant,con_,src_, 表示考场，考生，文件
# destination,dst_,cpy_, 表示目标，考生，文件
# 考场有份名单，从这份名单判断有没有提交成功
# 再两者相减，统计相差的人
# no_name, no_dir, no_file 没有该考生，没有子文件夹，没有文件
# 122行附近有sys.exit() 之后为未重构前的代码

from openpyxl import Workbook
from openpyxl import load_workbook
import os,shutil,sys
import numpy as np

wb0 = load_workbook("接口.xlsx")
ws0 = wb0.active
prob_list, extension_list = [], []

#####################################################

# contestant_dir = 'test'
# destination_dir = "all"
# room_name = "all"
# # contestant_dir = input()
# # destination_dir = input()
# # room_name = input()
# name_dir = 'name.xlsx'
# skipped = "考场"
# prob_list = ['number','work']
# extension_list = ['.cpp','.c','.pas']
# save_xlsx_name = "sample.xlsx"

contestant_dir = ws0.cell(row=2, column=2).value 
if not contestant_dir: contestant_dir = input()
destination_dir = ws0.cell(row=3, column=2).value 
if not destination_dir: destination_dir = input()
room_name = ws0.cell(row=4, column=2).value 
if not room_name: room_name = input()
name_dir = ws0.cell(row=5, column=2).value 
if not name_dir: name_dir = input()

skipped = ws0.cell(row=6, column=2).value 
if not skipped: skipped = input()

for i in range(2,10):
    if not ws0.cell(row=7, column=i).value: break 
    prob_list.append(ws0.cell(row=8, column=i).value)
 
for i in range(2,10):
    if not ws0.cell(row=8, column=i).value: break 
    extension_list.append(ws0.cell(row=8, column=i).value)

save_xlsx_name = ws0.cell(row=9, column=2).value 
if not save_xlsx_name: save_xlsx_name = input()

# contestant_dir是考生文件夹路径,destination_dir是目标文件夹路径
# room_name是考场名称 name_dir是考生名单名字 skipped 跳过第一行
# prob_list是题目名字 extension_list是认可后缀
# save_xlsx_name是保存的位置（格式为xlsx）

# 修改区域
#####################################################

if contestant_dir and destination_dir and room_name and name_dir and skipped and prob_list and extension_list and save_xlsx_name:
    print("接口正确")
else :
    print("接口有误")
    sys.exit()

name_list, room_list = [], []
no_name, no_dir, no_file = [], [], []

wb = Workbook()
ws = wb.active
wb2 = load_workbook(name_dir)
namewb = wb2.active
max_row = namewb.max_row
print("the number of contestant is %d"%max_row)

for row in range(1,max_row):
    name_list.append(namewb.cell(row=row,column=1).value)
    room_list.append(namewb.cell(row=row,column=2).value)
# print(name_list)

ws.cell(row=1,column=1,value="stuID")
ws.cell(row=1,column=2,value="room")
ws.cell(row=1,column=3,value="status")
idx=2
for i, con_name in enumerate(name_list):
    con_room = room_list[i]
    if con_room == skipped : continue
    if con_room != room_name and room_name != "all": continue
    con_path = os.path.join(contestant_dir, con_name)
    
    if not os.path.exists(con_path):
        no_name.append(con_name)
        print("%s in %s is absent"%(con_name,con_room))
        ws.cell(row=idx,column=1,value=con_name)
        ws.cell(row=idx,column=2,value=con_room)
        ws.cell(row=idx,column=3,value="abs_con")
        idx=idx+1
        continue
    if destination_dir != '0':
        dst_path = os.path.join(destination_dir,con_name)
        if not os.path.exists(dst_path):
            os.makedirs(dst_path)

    prob_found, file_found = [], []
    for p_name in prob_list:
        src_path = os.path.join(con_path, p_name)
        if not os.path.exists(src_path): continue
        pro_file = set(os.path.join(src_path,p_name+ext) for ext in extension_list)
        src_file = set(pro_file) & set(os.path.join(src_path,x) for x in os.listdir(src_path))
        prob_found.append(src_path) 
        if src_file:
            src_file = list(src_file)
            aim_file = src_file[0]
            tmp, file_name = os.path.split(aim_file) 
            
            file_found.append(os.path.join(src_path, aim_file))
            
            if destination_dir != '0':
                cpy_path = os.path.join(dst_path, p_name)
                if not os.path.exists(cpy_path):
                    os.makedirs(cpy_path)
                cpy_file = os.path.join(cpy_path,file_name)
                shutil.copy2(aim_file,cpy_file)
        
    if not prob_found:
        no_dir.append(con_name) # 没有子文件夹
        print("%s in %s have no sub_fold"%(con_name,room_name))
        ws.cell(row=idx,column=1,value=con_name)
        ws.cell(row=idx,column=2,value=con_room)
        ws.cell(row=idx,column=3,value="no_dir")
        idx+=1

    else :
        if not file_found:
            no_file.append(con_name) # 没有程序
            print("%s in %s have no file"%(con_name,room_name))
            ws.cell(row=idx,column=1,value=con_name)
            ws.cell(row=idx,column=2,value=con_room)
            ws.cell(row=idx,column=3,value="no_file")
            idx+=1
        else :
            ws.cell(row=idx,column=1,value=con_name)
            ws.cell(row=idx,column=2,value=con_room)
            ws.cell(row=idx,column=3,value="Succ")
            idx+=1

redundant_con = list (set(os.listdir(contestant_dir)) - set(name_list))
redundant_con = sorted(redundant_con)
for name in redundant_con:
    print("%s in none is redudant"%name)
    ws.cell(row=idx,column=1,value=name)
    ws.cell(row=idx,column=2,value="none")
    ws.cell(row=idx,column=3,value="red_con")
    idx=idx+1 

wb.save(save_xlsx_name)

sys.exit()





    # if 'GD-00004' in col:
    #     print("TES")

for i in range(1,max_row):
    stuid=namewb.cell(row=i, column=1).value
    studstpath=dstdir+stuid
    # print(studstpath)
    # print(namewb.cell(row=i,column=1).value)
    if not os.path.exists(studstpath):
        print("%s has not submit successfully"%stuid)

    


srcdir='J1/'
dstdir='test1/'
# srcdir=input()
# dstdir=input()
namelist='name.xlsx'
prob=['number','work']
extension=['.cpp','.pas']


wb = Workbook()
ws = wb.active


# srcfile='/Users/sky48/Desktop/python/test'
# dstfile='/Users/sky48/Desktop/python/test1/'


for i,element in enumerate(prob):
    ws.cell ( row=1, column=2+i, value=element)

size=len(prob)
print("the number of problem is %d"%size)
idx=1;
# for col in range(0,size):
    # ws.cell (row=1, column=2+col, value=prob[col])
ws.cell (row=1, column=1,value="stuID")
ws.cell (row=1, column=2+size,value="successful")

filelist=os.listdir(srcdir)
dstlist=os.listdir(dstdir)
for idx,files in enumerate(filelist): # files is GD-00000
    # fpath,fname=os.path.split(files)
    idx=idx+2;
    stupath=os.path.join(srcdir,files)
    stulist=os.listdir(stupath)  # 到GD-00000里了
    issubmit=False
    ws.cell(row=idx,column=1,value=files)
    for i,prog in enumerate(prob):
        # cols=cols+1
        havefile=0;
        for j,ext in enumerate(extension):
            anspath=stupath+'/'+prog+'/'+prog+ext
            dstpath=dstdir+files+'/'+prog+'/'+prog+ext
            dstdirpath=dstdir+files+'/'+prog
            # print(dstpath)
            # print(anspath)
            if os.path.exists(anspath):
                ws.cell(row=idx,column=i+2,value=ext)
                havefile=1 
                issubmit=1
                if not os.path.exists(dstdirpath):
                    
                    os.makedirs(dstdirpath)       #创建路径
                shutil.copy2(anspath,dstpath)
                break
        if(havefile==0):
            ws.cell(row=idx,column=i+2,value="NONE")
        

    if(issubmit):
        ws.cell(row=idx,column=size+2,value="Succ")
    else :
        ws.cell(row=idx,column=size+2,value="Fail")
        print("%s has no files"%files)

        # fext,fname=os.path.splitext(prog)
        # print(files)
        # print(prog)
        # print(fname)
        # print(fext)
        # if(fext != 'cpp'): 
        #     continue

        # srcpath = srcdir + files 
        # dstpath = dstdir + files 
        # shutil.move(srcpath,dstpath)
        # print(files)
        # print(srcpath)


# Data can be assigned directly to cells

    # prob["cell"]

# ws['A1'] = 42

# # Rows can also be appended
# ws.append([1, 2, 3])

wb.save("sample.xlsx")
