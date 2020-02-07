# contestant,con_,src_, 表示考场，考生，文件
# destination,dst_,cpy_, 表示目标，考生，文件
# 每个考场有份名单，从这份名单判断有没有提交成功
# 再两者相减，统计相差的人
# no_name, no_dir, no file 没有该考生，没有子文件夹，没有文件


from openpyxl import Workbook
from openpyxl import load_workbook
import os,shutil,sys
import numpy as np

# extension_list = ['cpp','c','pas']
# p_name = "23" 
# a = set([p_name + ext for ext in extension_list])
# print(a)

#####################################################

# contestant_dir = 'test'
# destination_dir = 'all'
contestant_dir = input()
destination_dir = input()
name_dir = 'name.xlsx'
prob_list = ['number','work']
extension_list = ['.cpp','.c','.pas']

# 修改区域
#####################################################

name_list = []
no_name, no_dir, no_file = [], [], []

wb = Workbook()
ws = wb.active
wb2 = load_workbook(name_dir)
namewb = wb2.active
max_row = namewb.max_row
print("the number of contestant is %d"%max_row)

for row in range(1,max_row):
    name_list.append(namewb.cell(row=row,column=1).value)
# print(name_list)

ws.cell(row=1,column=1,value="stuID")
ws.cell(row=1,column=2,value="stated")
idx=2
for con_name in name_list:
    con_path = os.path.join(contestant_dir, con_name)
    if not os.path.exists(con_path):
        no_name.append(con_name)
        continue
    dst_path = os.path.join(destination_dir,con_name)
    if not os.path.exists(dst_path):
       os.makedirs(dst_path)

    prob_found, file_found = [], []
    for p_name in prob_list:
        src_path = os.path.join(con_path, p_name)
        if not os.path.exists(src_path):
            continue;
        
        pro_file = set(os.path.join(src_path,p_name+ext) for ext in extension_list)
        src_file = set(pro_file) & set(os.path.join(src_path,x) for x in os.listdir(src_path))
        prob_found.append(src_path) 
        if src_file:
            src_file = list(src_file)
            aim_file = src_file[0]
            tmp, file_name = os.path.split(aim_file) 
            
            file_found.append(os.path.join(src_path, aim_file))
            cpy_path = os.path.join(dst_path, p_name)
            if not os.path.exists(cpy_path):
                os.makedirs(cpy_path)
            cpy_file = os.path.join(cpy_path,file_name)
            shutil.copy2(aim_file,cpy_file)
        
    if not prob_found:
        no_dir.append(con_name) # 没有子文件夹
    else :
        if not file_found:
            no_file.append(con_name) # 没有程序
        else :
            ws.cell(row=idx,column=1,value=con_name)
            ws.cell(row=idx,column=2,value="Succ")
            idx=idx+1

absent_con = list (set(name_list) - set(os.listdir(contestant_dir)))
redundant_con = list (set(os.listdir(contestant_dir)) - set(name_list))

no_dir = sorted(no_dir)
no_file = sorted(no_file)
absent_con = sorted(absent_con)
redundant_con = sorted(redundant_con)

for name in no_dir:
    print("%s have no sub_fold"%name)
    ws.cell(row=idx,column=1,value=name)
    ws.cell(row=idx,column=2,value="no_dir")
    idx=idx+1

for name in no_file:
    print("%s have no file"%name)
    ws.cell(row=idx,column=1,value=name)
    ws.cell(row=idx,column=2,value="no_file")
    idx+=1

for name in absent_con:
    print("%s is absent"%name)
    ws.cell(row=idx,column=1,value=name)
    ws.cell(row=idx,column=2,value="abs_con")
    idx=idx+1

for name in redundant_con:
    print("%s is redudant"%name)
    ws.cell(row=idx,column=1,value=name)
    ws.cell(row=idx,column=2,value="red_con")
    idx=idx+1        

wb.save("sample.xlsx")



sys.exit()


    # if 'GD-00004' in col:
    #     print("TES")

for i in range(1,max_row):
    stuid=namewb.cell(row=i, column=1).value
    studstpath=dstdir+stuid
    # print(studstpath)
    # print(namewb.cell(row=i,column=1).value)
    if not os.path.exists(studstpath):
        print("%s has not submit successfully"%stuid)

    


srcdir='J1/'
dstdir='test1/'
# srcdir=input()
# dstdir=input()
namelist='name.xlsx'
prob=['number','work']
extension=['.cpp','.pas']


wb = Workbook()
ws = wb.active


# srcfile='/Users/sky48/Desktop/python/test'
# dstfile='/Users/sky48/Desktop/python/test1/'


for i,element in enumerate(prob):
    ws.cell ( row=1, column=2+i, value=element)

size=len(prob)
print("the number of problem is %d"%size)
idx=1;
# for col in range(0,size):
    # ws.cell (row=1, column=2+col, value=prob[col])
ws.cell (row=1, column=1,value="stuID")
ws.cell (row=1, column=2+size,value="successful")

filelist=os.listdir(srcdir)
dstlist=os.listdir(dstdir)
for idx,files in enumerate(filelist): # files is GD-00000
    # fpath,fname=os.path.split(files)
    idx=idx+2;
    stupath=os.path.join(srcdir,files)
    stulist=os.listdir(stupath)  # 到GD-00000里了
    issubmit=False
    ws.cell(row=idx,column=1,value=files)
    for i,prog in enumerate(prob):
        # cols=cols+1
        havefile=0;
        for j,ext in enumerate(extension):
            anspath=stupath+'/'+prog+'/'+prog+ext
            dstpath=dstdir+files+'/'+prog+'/'+prog+ext
            dstdirpath=dstdir+files+'/'+prog
            # print(dstpath)
            # print(anspath)
            if os.path.exists(anspath):
                ws.cell(row=idx,column=i+2,value=ext)
                havefile=1 
                issubmit=1
                if not os.path.exists(dstdirpath):
                    
                    os.makedirs(dstdirpath)       #创建路径
                shutil.copy2(anspath,dstpath)
                break
        if(havefile==0):
            ws.cell(row=idx,column=i+2,value="NONE")
        

    if(issubmit):
        ws.cell(row=idx,column=size+2,value="Succ")
    else :
        ws.cell(row=idx,column=size+2,value="Fail")
        print("%s has no files"%files)

        # fext,fname=os.path.splitext(prog)
        # print(files)
        # print(prog)
        # print(fname)
        # print(fext)
        # if(fext != 'cpp'): 
        #     continue

        # srcpath = srcdir + files 
        # dstpath = dstdir + files 
        # shutil.move(srcpath,dstpath)
        # print(files)
        # print(srcpath)


# Data can be assigned directly to cells

    # prob["cell"]

# ws['A1'] = 42

# # Rows can also be appended
# ws.append([1, 2, 3])

wb.save("sample.xlsx")
