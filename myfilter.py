
# def mymovefile(srcfile,dstfile):
#     if not os.path.isfile(srcfile):
#         print ("%s not exist!"%(srcfile))
#     else:
#         fpath,fname=os.path.split(dstfile)    #分离文件名和路径
#         if not os.path.exists(fpath):
#             os.makedirs(fpath)                #创建路径
#         shutil.move(srcfile,dstfile)          #移动文件
#         print ("move %s -> %s"%( srcfile,dstfile))

# def mycopyfile(srcfile,dstfile):
#     if not os.path.isfile(srcfile):
#         print( "%s not exist!"%(srcfile))
#     else:
#         fpath,fname=os.path.split(dstfile)    #分离文件名和路径
#         if not os.path.exists(fpath):
#             os.makedirs(fpath)                #创建路径
#         shutil.copyfile(srcfile,dstfile)      #复制文件
#         print( "copy %s -> %s"%( srcfile,dstfile))


from openpyxl import Workbook
from openpyxl import load_workbook
import os,shutil
import numpy as np

srcdir='J1/'
dstdir='test1/'
# srcdir=input()
# dstdir=input()
namelist='name.xlsx'
prob=['number','work']
extension=['.cpp','.pas']


wb = Workbook()
ws = wb.active


# srcfile='/Users/sky48/Desktop/python/test'
# dstfile='/Users/sky48/Desktop/python/test1/'


for i,element in enumerate(prob):
    ws.cell ( row=1, column=2+i, value=element)

size=len(prob)
print("the number of problem is %d"%size)
idx=1;
# for col in range(0,size):
    # ws.cell (row=1, column=2+col, value=prob[col])
ws.cell (row=1, column=1,value="stuID")
ws.cell (row=1, column=2+size,value="successful")

filelist=os.listdir(srcdir)
dstlist=os.listdir(dstdir)
for idx,files in enumerate(filelist): # files is GD-00000
    # fpath,fname=os.path.split(files)
    idx=idx+2;
    stupath=os.path.join(srcdir,files)
    stulist=os.listdir(stupath)  # 到GD-00000里了
    issubmit=False
    ws.cell(row=idx,column=1,value=files)
    for i,prog in enumerate(prob):
        # cols=cols+1
        havefile=0;
        for j,ext in enumerate(extension):
            anspath=stupath+'/'+prog+'/'+prog+ext
            dstpath=dstdir+files+'/'+prog+'/'+prog+ext
            dstdirpath=dstdir+files+'/'+prog
            # print(dstpath)
            # print(anspath)
            if os.path.exists(anspath):
                ws.cell(row=idx,column=i+2,value=ext)
                havefile=1 
                issubmit=1
                if not os.path.exists(dstdirpath):
                    
                    os.makedirs(dstdirpath)       #创建路径
                shutil.copy2(anspath,dstpath)
                break
        if(havefile==0):
            ws.cell(row=idx,column=i+2,value="NONE")
        

    if(issubmit):
        ws.cell(row=idx,column=size+2,value="Succ")
    else :
        ws.cell(row=idx,column=size+2,value="Fail")
        print("%s has no files"%files)

        # fext,fname=os.path.splitext(prog)
        # print(files)
        # print(prog)
        # print(fname)
        # print(fext)
        # if(fext != 'cpp'): 
        #     continue

        # srcpath = srcdir + files 
        # dstpath = dstdir + files 
        # shutil.move(srcpath,dstpath)
        # print(files)
        # print(srcpath)


# Data can be assigned directly to cells

    # prob["cell"]

# ws['A1'] = 42

# # Rows can also be appended
# ws.append([1, 2, 3])

wb.save("sample.xlsx")

wb2 = load_workbook(namelist)
namewb = wb2.active

max_row=namewb.max_row
print("the max row is %s"%max_row)
    
for col in namewb.iter_cols(min_row=1,max_col=1,values_only=True):
    print(col)
    if 'GD-00004' in col:
        print("TES")
    


for i in range(1,max_row):
    stuid=namewb.cell(row=i, column=1).value
    studstpath=dstdir+stuid
    # print(studstpath)
    # print(namewb.cell(row=i,column=1).value)
    if not os.path.exists(studstpath):
        print("%s has not submit successfully"%stuid)

    
