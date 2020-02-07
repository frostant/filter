
# def mymovefile(srcfile,dstfile):
#     if not os.path.isfile(srcfile):
#         print ("%s not exist!"%(srcfile))
#     else:
#         fpath,fname=os.path.split(dstfile)    #分离文件名和路径
#         if not os.path.exists(fpath):
#             os.makedirs(fpath)                #创建路径
#         shutil.move(srcfile,dstfile)          #移动文件
#         print ("move %s -> %s"%( srcfile,dstfile))

# def mycopyfile(srcfile,dstfile):
#     if not os.path.isfile(srcfile):
#         print( "%s not exist!"%(srcfile))
#     else:
#         fpath,fname=os.path.split(dstfile)    #分离文件名和路径
#         if not os.path.exists(fpath):
#             os.makedirs(fpath)                #创建路径
#         shutil.copyfile(srcfile,dstfile)      #复制文件
#         print( "copy %s -> %s"%( srcfile,dstfile))


from openpyxl import Workbook
import os,shutil
import numpy as np
# srcfile='tt.txt'
srcdir='J1/'
# dstfile='test/tt.txt'
dstdir='test1/'

wb = Workbook()
ws = wb.active

# srcdir=input()
# dstdir=input()

# srcfile='/Users/sky48/Desktop/python/test'
# dstfile='/Users/sky48/Desktop/python/test1/'
prob=['number','work']
size=len(prob)
print("size=%d"%size)
idx=1;
for col in range(0,size):
    ws.cell (row=1, column=2+col, value=prob[col])
ws.cell (row=1, column=1,value="stuID")
ws.cell (row=1, column=2+size,value="successful")

filelist=os.listdir(srcdir)
for files in filelist: # files is GD-00000
    # fpath,fname=os.path.split(files)
    idx=idx+1; 
    print(files)
    stupath=os.path.join(srcdir,files)
    stulist=os.listdir(stupath)  # 到GD-00000里了
    cols=1
    isfile=0
    ws.cell(row=idx,column=1,value=files)
    for prog in prob:
        cols=cols+1
        anspath=stupath+'/'+prog+'/'+prog+'.cpp'
        dstpath=dstdir+files+'/'+prog+'/'+prog+'.cpp'
        dstdirpath=dstdir+files+'/'+prog
        # print(dstpath)
        # print(anspath)
        if os.path.exists(anspath):
            ws.cell(row=idx,column=cols,value="cpp")
            isfile=1
        else:
            ws.cell(row=idx,column=cols,value="NONE")
        if not os.path.exists(anspath):
            continue
        if not os.path.exists(dstdirpath):
            os.makedirs(dstdirpath)       #创建路径
        shutil.copyfile(anspath,dstpath)
    
    cols=cols+1
    if(isfile):
        ws.cell(row=idx,column=cols,value="Succ")
    else :
        ws.cell(row=idx,column=cols,value="Fail")

        # fext,fname=os.path.splitext(prog)
        # print(files)
        # print(prog)
        # print(fname)
        # print(fext)
        # if(fext != 'cpp'): 
        #     continue

        # srcpath = srcdir + files 
        # dstpath = dstdir + files 
        # shutil.move(srcpath,dstpath)
        # print(files)
        # print(srcpath)


# Data can be assigned directly to cells

    # prob["cell"]

# ws['A1'] = 42

# # Rows can also be appended
# ws.append([1, 2, 3])

wb.save("sample.xlsx")
